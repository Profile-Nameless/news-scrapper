from googlesearch import search
from bs4 import BeautifulSoup
import requests
import openpyxl as pxl
import re



def newssearch(query,language="en"):
    result=search(query,lang=language,num=10,stop=10,extra_params={'tbm': 'nws'})
    data=[]
    for i in result:
        response=requests.get(i)

        if response.status_code==200:
            content=response.content
            soup=BeautifulSoup(content,'html.parser')
            title=soup.title.string
            description_tag = soup.find('meta', attrs={'name': 'og:description'}) or soup.find('meta', attrs={'property': 'og:description'})

            if description_tag and 'content' in description_tag.attrs:
                description = description_tag['content']
            else:
                description = 'No description available'
            publisher_tag = soup.select_one(".correct-publisher-selector")


# To get the author's name
            author_tag=soup.find(class_=re.compile("author"))
            if author_tag:
                author=author_tag.get_text(strip=True)
            else:
                author=None
            # To get the publishing date
            date_tag = soup.select_one(".correct-date-selector")
            if date_tag:
                date = date_tag.get_text(strip=True)
            else:
                date = 'No date available'
            data.append([i,title, description,author,date])
    print(author_tag)
    return data
        
def createxl(newsquery):
    data=newssearch(newsquery)
    try:   
        wb=pxl.load_workbook('newslinks.xlsx')
        try:
            ws=wb[newsquery+'newslink']
        except:
            ws=wb.create_sheet(newsquery+'newslink')
            ws.append(['URL','Title','Description','Author','Publishing Date'])
    
    except:
        
        wb=pxl.Workbook()
        ws=wb.active
        ws.title=newsquery+'newslink'
        ws.append(['URL','Title','Description','Author','Publishing Date'])
    
    finally:    
        
        for row in data:
                    ws.append(row)
        wb.save('newslinks.xlsx')

newsquery=input("Enter search query:")
createxl(newsquery)
